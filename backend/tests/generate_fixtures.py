from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook


FIXTURE_DIR = Path(__file__).resolve().parent / "fixtures"


def _write_csv(path: Path, headers: list[str], rows: list[list[object]]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(headers)
        writer.writerows(rows)


def _save_workbook(path: Path, sheets: dict[str, tuple[list[str], list[list[object]]]]) -> None:
    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)
    for sheet_name, (headers, rows) in sheets.items():
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(headers)
        for row in rows:
            sheet.append(row)
    workbook.save(path)


def _save_xls(path: Path, headers: list[str], rows: list[list[object]]) -> None:
    try:
        import xlwt
    except ModuleNotFoundError:
        return

    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet("Sheet1")
    for column_index, header in enumerate(headers):
        sheet.write(0, column_index, header)
    for row_index, row in enumerate(rows, start=1):
        for column_index, value in enumerate(row):
            sheet.write(row_index, column_index, value)
    workbook.save(str(path))


def ensure_fixtures() -> None:
    FIXTURE_DIR.mkdir(parents=True, exist_ok=True)

    daily_rows = [
        ["2026-06-01", 12000],
        ["2026/06/02", 12500],
        ["20260603", 13000],
    ] + [[f"2026-06-{day:02d}", 13000 + day * 100] for day in range(4, 31)]
    _write_csv(FIXTURE_DIR / "daily_air_passengers.csv", ["date", "passenger_count"], daily_rows)

    monthly_rows = [
        ["2026-01", 120000],
        ["2026/02", 125000],
        ["2026\u5e743\u6708", 130000],
    ] + [[f"2026-{month:02d}", 130000 + month * 1000] for month in range(4, 13)]
    _save_workbook(FIXTURE_DIR / "monthly_air_passengers.xlsx", {"monthly": (["month", "passenger_count"], monthly_rows)})

    domestic_rows: list[list[object]] = []
    international_rows: list[list[object]] = []
    for day in range(1, 31):
        domestic_rows.append([f"CZ{day:03d}", f"2026-06-{day:02d}", "CAN", 180 + day])
        domestic_rows.append([f"MU{day:03d}", f"2026-06-{day:02d}", "SHA", 210 + day])
        international_rows.append([f"CA{day:03d}", f"2026/06/{day:02d}", "PEK", 80 + day])
        international_rows.append([f"HU{day:03d}", f"202606{day:02d}", "SZX", 95 + day])
    _save_workbook(
        FIXTURE_DIR / "raw_flight_detail_multi_sheet.xlsx",
        {
            "domestic": (["flight_no", "flight_date", "airport", "passenger_count"], domestic_rows),
            "international": (["flight_no", "flight_date", "airport", "passenger_count"], international_rows),
        },
    )

    _write_csv(FIXTURE_DIR / "invalid_date.csv", ["date", "passenger_count"], [["not-a-date", 1], ["also-bad", 2]])
    _save_workbook(
        FIXTURE_DIR / "duplicate_dates.xlsx",
        {"duplicates": (["date", "passenger_count"], [["2026-06-01", 100], ["2026-06-01", 120], ["2026-06-02", 130]])},
    )
    _save_workbook(
        FIXTURE_DIR / "missing_values.xlsx",
        {"missing": (["date", "passenger_count"], [["2026-06-01", 100], ["2026-06-03", 130], ["2026-06-04", None]])},
    )
    _write_csv(FIXTURE_DIR / "short_series.csv", ["date", "passenger_count"], [["2026-06-01", 10], ["2026-06-02", 20]])
    _save_xls(FIXTURE_DIR / "legacy_daily_air_passengers.xls", ["date", "passenger_count"], daily_rows)


if __name__ == "__main__":
    ensure_fixtures()
