from __future__ import annotations

import csv
from datetime import date, datetime
from pathlib import Path
from typing import Any

import pandas as pd

from app.core.errors import AppError
from app.core.storage import get_upload_path, read_upload_metadata
from app.schemas import SheetPreview, UploadPreviewResponse
from app.services.schema_profiler import profile_columns


CSV_SHEET_NAME = "CSV"
CSV_SAMPLE_CHARS = 1024 * 1024
CSV_DELIMITERS = [",", "\t", ";", "|"]


def _clean_header(values: list[Any]) -> list[str]:
    headers: list[str] = []
    seen: dict[str, int] = {}
    for index, value in enumerate(values):
        name = str(value).strip() if value is not None and str(value).strip() else f"column_{index + 1}"
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 1
        headers.append(name)
    return headers


def _json_value(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if pd.isna(value):
        return None
    return value


def _fallback_csv_dialect(delimiter: str):
    return type("FallbackCsvDialect", (csv.excel,), {"delimiter": delimiter})


def _detect_csv_dialect(sample: str):
    if not sample:
        return csv.excel
    try:
        return csv.Sniffer().sniff(sample, delimiters="".join(CSV_DELIMITERS))
    except csv.Error:
        first_line = next((line for line in sample.splitlines() if line.strip()), sample)
        delimiter = max(CSV_DELIMITERS, key=lambda item: first_line.count(item))
        if first_line.count(delimiter) == 0:
            delimiter = ","
        return _fallback_csv_dialect(delimiter)


def _read_csv_sample(path: Path, encoding: str) -> str:
    with path.open("r", encoding=encoding, newline="") as handle:
        return handle.read(CSV_SAMPLE_CHARS)


def _csv_read_options(path: Path) -> dict[str, str]:
    for encoding in ("utf-8-sig", "gb18030"):
        try:
            sample = _read_csv_sample(path, encoding)
        except UnicodeDecodeError:
            continue
        dialect = _detect_csv_dialect(sample)
        return {"encoding": encoding, "sep": dialect.delimiter}
    raise AppError("Failed to decode the CSV file. Please use UTF-8 or GB18030 encoding.")


def _rows_from_records(records: list[list[Any]], headers: list[str], limit: int) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for record in records[:limit]:
        row = {}
        for index, header in enumerate(headers):
            row[header] = _json_value(record[index]) if index < len(record) else None
        rows.append(row)
    return rows


def _preview_csv(path: Path, limit: int) -> SheetPreview:
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            sample = handle.read(CSV_SAMPLE_CHARS)
            handle.seek(0)
            dialect = _detect_csv_dialect(sample)
            reader = csv.reader(handle, dialect)
            try:
                headers = _clean_header(next(reader))
            except StopIteration as exc:
                raise AppError("The CSV file is empty.") from exc
            preview_records: list[list[Any]] = []
            row_count = 0
            for record in reader:
                row_count += 1
                if len(preview_records) < limit:
                    preview_records.append(record)
    except UnicodeDecodeError:
        with path.open("r", encoding="gb18030", newline="") as handle:
            sample = handle.read(CSV_SAMPLE_CHARS)
            handle.seek(0)
            reader = csv.reader(handle, _detect_csv_dialect(sample))
            try:
                headers = _clean_header(next(reader))
            except StopIteration as exc:
                raise AppError("The CSV file is empty.") from exc
            preview_records = []
            row_count = 0
            for record in reader:
                row_count += 1
                if len(preview_records) < limit:
                    preview_records.append(record)

    rows = _rows_from_records(preview_records, headers, limit)
    if not headers:
        raise AppError("No header row was found in the CSV file.")
    return SheetPreview(
        sheetName=CSV_SHEET_NAME,
        rowCountApprox=row_count,
        columns=profile_columns(rows, headers),
        previewRows=rows,
    )


def _preview_xlsx_sheet(path: Path, sheet_name: str, limit: int) -> SheetPreview:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise AppError("openpyxl is required to read xlsx files.") from exc

    workbook = load_workbook(filename=path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        workbook.close()
        raise AppError(f"Sheet '{sheet_name}' was not found.", 404)
    sheet = workbook[sheet_name]
    iterator = sheet.iter_rows(values_only=True)
    try:
        headers = _clean_header(list(next(iterator)))
    except StopIteration as exc:
        workbook.close()
        raise AppError(f"Sheet '{sheet_name}' is empty.") from exc

    records: list[list[Any]] = []
    row_count = 0
    for row in iterator:
        row_count += 1
        if len(records) < limit:
            records.append(list(row))
    rows = _rows_from_records(records, headers, limit)
    preview = SheetPreview(
        sheetName=sheet_name,
        rowCountApprox=row_count,
        columns=profile_columns(rows, headers),
        previewRows=rows,
    )
    workbook.close()
    return preview


def _preview_xls_sheet(path: Path, sheet_name: str, limit: int) -> SheetPreview:
    try:
        import xlrd
    except ImportError as exc:
        raise AppError("xlrd is required to read xls files.") from exc

    workbook = xlrd.open_workbook(path)
    if sheet_name not in workbook.sheet_names():
        raise AppError(f"Sheet '{sheet_name}' was not found.", 404)
    sheet = workbook.sheet_by_name(sheet_name)
    if sheet.nrows == 0:
        raise AppError(f"Sheet '{sheet_name}' is empty.")
    headers = _clean_header(sheet.row_values(0))
    records = [sheet.row_values(row_index) for row_index in range(1, min(sheet.nrows, limit + 1))]
    rows = _rows_from_records(records, headers, limit)
    return SheetPreview(
        sheetName=sheet_name,
        rowCountApprox=max(0, sheet.nrows - 1),
        columns=profile_columns(rows, headers),
        previewRows=rows,
    )


def preview_upload(upload_id: str, limit: int = 100) -> UploadPreviewResponse:
    metadata = read_upload_metadata(upload_id)
    path = get_upload_path(upload_id)
    ext = path.suffix.lower()

    if ext == ".csv":
        sheets = [_preview_csv(path, limit)]
    elif ext == ".xlsx":
        from openpyxl import load_workbook

        workbook = load_workbook(filename=path, read_only=True, data_only=True)
        sheet_names = workbook.sheetnames
        workbook.close()
        sheets = [_preview_xlsx_sheet(path, sheet_name, limit) for sheet_name in sheet_names]
    elif ext == ".xls":
        import xlrd

        workbook = xlrd.open_workbook(path, on_demand=True)
        sheet_names = workbook.sheet_names()
        workbook.release_resources()
        sheets = [_preview_xls_sheet(path, sheet_name, limit) for sheet_name in sheet_names]
    else:
        raise AppError("Unsupported file format. Please upload a csv, xlsx, or xls file.")

    return UploadPreviewResponse(
        uploadId=upload_id,
        fileName=metadata["fileName"],
        fileSize=metadata["fileSize"],
        fileSha256=metadata["fileSha256"],
        sheets=sheets,
    )


def preview_sheet(upload_id: str, sheet_name: str, limit: int = 100) -> SheetPreview:
    path = get_upload_path(upload_id)
    ext = path.suffix.lower()
    if ext == ".csv":
        return _preview_csv(path, limit)
    if ext == ".xlsx":
        return _preview_xlsx_sheet(path, sheet_name, limit)
    if ext == ".xls":
        return _preview_xls_sheet(path, sheet_name, limit)
    raise AppError("Unsupported file format. Please upload a csv, xlsx, or xls file.")


def read_sheet_dataframe(upload_id: str, sheet_name: str) -> pd.DataFrame:
    path = get_upload_path(upload_id)
    ext = path.suffix.lower()
    try:
        if ext == ".csv":
            return pd.read_csv(path, **_csv_read_options(path))
        if ext == ".xlsx":
            return pd.read_excel(path, sheet_name=sheet_name, engine="openpyxl")
        if ext == ".xls":
            return pd.read_excel(path, sheet_name=sheet_name, engine="xlrd")
    except Exception as exc:
        raise AppError(f"Failed to read the selected sheet: {exc}") from exc
    raise AppError("Unsupported file format. Please upload a csv, xlsx, or xls file.")
